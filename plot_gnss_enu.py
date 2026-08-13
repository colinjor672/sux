from __future__ import annotations

import argparse
import math
from pathlib import Path

from openpyxl import Workbook, load_workbook
from PIL import Image, ImageDraw, ImageFont


WGS84_A = 6378137.0
WGS84_F = 1.0 / 298.257223563
WGS84_E2 = WGS84_F * (2.0 - WGS84_F)


def geodetic_to_ecef(lon_deg: float, lat_deg: float, height: float) -> tuple[float, float, float]:
    lon = math.radians(lon_deg)
    lat = math.radians(lat_deg)
    sin_lat = math.sin(lat)
    cos_lat = math.cos(lat)
    radius = WGS84_A / math.sqrt(1.0 - WGS84_E2 * sin_lat * sin_lat)
    x = (radius + height) * cos_lat * math.cos(lon)
    y = (radius + height) * cos_lat * math.sin(lon)
    z = (radius * (1.0 - WGS84_E2) + height) * sin_lat
    return x, y, z


def make_enu_converter(lon0_deg: float, lat0_deg: float, height0: float):
    x0, y0, z0 = geodetic_to_ecef(lon0_deg, lat0_deg, height0)
    lon0 = math.radians(lon0_deg)
    lat0 = math.radians(lat0_deg)
    sin_lon, cos_lon = math.sin(lon0), math.cos(lon0)
    sin_lat, cos_lat = math.sin(lat0), math.cos(lat0)

    def convert(lon_deg: float, lat_deg: float, height: float) -> tuple[float, float, float]:
        x, y, z = geodetic_to_ecef(lon_deg, lat_deg, height)
        dx, dy, dz = x - x0, y - y0, z - z0
        east = -sin_lon * dx + cos_lon * dy
        north = -sin_lat * cos_lon * dx - sin_lat * sin_lon * dy + cos_lat * dz
        up = cos_lat * cos_lon * dx + cos_lat * sin_lon * dy + sin_lat * dz
        return east, north, up

    return convert


def is_number(value: object) -> bool:
    return isinstance(value, (int, float)) and math.isfinite(value)


def load_font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    windows_fonts = Path("C:/Windows/Fonts")
    candidates = ["msyhbd.ttc", "msyh.ttc"] if bold else ["msyh.ttc", "simhei.ttf"]
    for candidate in candidates:
        font_path = windows_fonts / candidate
        if font_path.exists():
            return ImageFont.truetype(str(font_path), size=size)
    return ImageFont.load_default()


def nice_ticks(low: float, high: float, target: int = 7) -> list[float]:
    span = max(high - low, 1e-9)
    raw_step = span / target
    magnitude = 10 ** math.floor(math.log10(raw_step))
    step = min((1, 2, 5, 10), key=lambda value: abs(value * magnitude - raw_step)) * magnitude
    start = math.ceil(low / step) * step
    ticks = []
    value = start
    while value <= high + step * 1e-8:
        ticks.append(value)
        value += step
    return ticks


def convert_workbook(input_path: Path, output_xlsx: Path):
    source = load_workbook(input_path, read_only=True, data_only=True)
    sheet = source.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    while headers and headers[-1] is None:
        headers.pop()
    header_index = {str(value).strip(): index for index, value in enumerate(headers) if value is not None}
    required = ("Lon", "Lat", "Altitude (m)")
    missing = [name for name in required if name not in header_index]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    lon_i, lat_i, alt_i = (header_index[name] for name in required)
    first = next(
        (row for row in rows if all(is_number(row[i]) for i in (lon_i, lat_i, alt_i))),
        None,
    )
    if first is None:
        raise ValueError("No valid longitude, latitude, and altitude rows found")

    origin = (float(first[lon_i]), float(first[lat_i]), float(first[alt_i]))
    to_enu = make_enu_converter(*origin)
    converted: list[tuple[object, ...] | None] = []
    plot_east: list[float] = []
    plot_north: list[float] = []
    valid_enu: list[tuple[float, float, float]] = []

    output = Workbook()
    output_sheet = output.active
    output_sheet.title = "ENU相对坐标"
    output_headers = list(headers) + ["East (m)", "North (m)", "Up (m)"]
    output_sheet.append(output_headers)

    for row in rows:
        if all(i < len(row) and is_number(row[i]) for i in (lon_i, lat_i, alt_i)):
            enu = to_enu(float(row[lon_i]), float(row[lat_i]), float(row[alt_i]))
            converted.append(tuple(row[: len(headers)]) + enu)
            output_sheet.append(tuple(row[: len(headers)]) + enu)
            valid_enu.append(enu)
            plot_east.append(enu[0])
            plot_north.append(enu[1])
        else:
            converted.append(None)
            output_sheet.append(tuple(row[: len(headers)]) + (None, None, None))
            if plot_east and not math.isnan(plot_east[-1]):
                plot_east.append(math.nan)
                plot_north.append(math.nan)

    output_sheet.freeze_panes = "A2"
    for col in output_sheet.iter_cols(min_col=len(headers) + 1, max_col=len(headers) + 3):
        for cell in col[1:]:
            cell.number_format = "0.0000"
    output.save(output_xlsx)
    return origin, valid_enu, plot_east, plot_north


def plot_enu(
    output_png: Path,
    origin: tuple[float, float, float],
    valid_enu: list[tuple[float, float, float]],
    east: list[float],
    north: list[float],
) -> None:
    valid_east = [point[0] for point in valid_enu]
    valid_north = [point[1] for point in valid_enu]
    east_span = max(valid_east) - min(valid_east)
    north_span = max(valid_north) - min(valid_north)
    width, height = 1600, 1600
    image = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(image)
    title_font = load_font(42, bold=True)
    label_font = load_font(28)
    tick_font = load_font(22)
    note_font = load_font(21)
    legend_font = load_font(24)

    left, top, right, bottom = 170, 105, 1515, 1430
    plot_width, plot_height = right - left, bottom - top
    pad = max(east_span, north_span) * 0.06
    x_mid = (min(valid_east) + max(valid_east)) / 2
    y_mid = (min(valid_north) + max(valid_north)) / 2
    scale = min(plot_width / (east_span + 2 * pad), plot_height / (north_span + 2 * pad))
    visible_x_span = plot_width / scale
    visible_y_span = plot_height / scale
    x_min, x_max = x_mid - visible_x_span / 2, x_mid + visible_x_span / 2
    y_min, y_max = y_mid - visible_y_span / 2, y_mid + visible_y_span / 2

    def project(x: float, y: float) -> tuple[float, float]:
        return left + (x - x_min) * scale, bottom - (y - y_min) * scale

    for tick in nice_ticks(x_min, x_max):
        x, _ = project(tick, y_min)
        draw.line((x, top, x, bottom), fill="#dddddd", width=2)
        label = f"{tick:g}"
        box = draw.textbbox((0, 0), label, font=tick_font)
        draw.text((x - (box[2] - box[0]) / 2, bottom + 12), label, font=tick_font, fill="#202020")
    for tick in nice_ticks(y_min, y_max):
        _, y = project(x_min, tick)
        draw.line((left, y, right, y), fill="#dddddd", width=2)
        label = f"{tick:g}"
        box = draw.textbbox((0, 0), label, font=tick_font)
        draw.text((left - 16 - (box[2] - box[0]), y - (box[3] - box[1]) / 2), label, font=tick_font, fill="#202020")

    draw.rectangle((left, top, right, bottom), outline="#242424", width=3)
    segment: list[tuple[float, float]] = []
    for x, y in zip(east, north):
        if math.isnan(x) or math.isnan(y):
            if len(segment) >= 2:
                draw.line(segment, fill="#cf2f5b", width=4, joint="curve")
            segment = []
        else:
            segment.append(project(x, y))
    if len(segment) >= 2:
        draw.line(segment, fill="#cf2f5b", width=4, joint="curve")

    start_x, start_y = project(valid_east[0], valid_north[0])
    end_x, end_y = project(valid_east[-1], valid_north[-1])
    radius = 11
    draw.ellipse((start_x - radius, start_y - radius, start_x + radius, start_y + radius), fill="#15831c")
    cross = 13
    draw.line((end_x - cross, end_y - cross, end_x + cross, end_y + cross), fill="#e21a1a", width=6)
    draw.line((end_x - cross, end_y + cross, end_x + cross, end_y - cross), fill="#e21a1a", width=6)

    title = "GNSS轨迹的ENU相对坐标"
    title_box = draw.textbbox((0, 0), title, font=title_font)
    draw.text(((width - title_box[2]) / 2, 30), title, font=title_font, fill="#151515")
    x_label = "East (m)"
    x_box = draw.textbbox((0, 0), x_label, font=label_font)
    draw.text(((left + right - x_box[2]) / 2, 1490), x_label, font=label_font, fill="#151515")
    y_label = "North (m)"
    y_layer = Image.new("RGBA", (260, 60), (255, 255, 255, 0))
    ImageDraw.Draw(y_layer).text((0, 0), y_label, font=label_font, fill="#151515")
    y_layer = y_layer.rotate(90, expand=True)
    image.paste(y_layer, (35, (top + bottom - y_layer.height) // 2), y_layer)

    annotation = [
        f"East 范围: {east_span:.2f} m",
        f"North 范围: {north_span:.2f} m",
        f"有效点数: {len(valid_enu)}",
        f"ENU原点: {origin[0]:.7f}°, {origin[1]:.7f}°, {origin[2]:.3f} m",
    ]
    note_x, note_y = left + 18, top + 18
    note_width, note_height = 570, 132
    draw.rounded_rectangle(
        (note_x, note_y, note_x + note_width, note_y + note_height),
        radius=8,
        fill="#fff1cd",
        outline="#745d31",
        width=2,
    )
    for index, line in enumerate(annotation):
        draw.text((note_x + 12, note_y + 8 + index * 29), line, font=note_font, fill="#252018")

    legend_x, legend_y = right - 155, top + 18
    draw.rounded_rectangle((legend_x, legend_y, right - 18, legend_y + 78), radius=7, fill="white", outline="#b8b8b8", width=2)
    draw.ellipse((legend_x + 12, legend_y + 12, legend_x + 32, legend_y + 32), fill="#15831c")
    draw.text((legend_x + 43, legend_y + 7), "起点", font=legend_font, fill="#202020")
    draw.line((legend_x + 12, legend_y + 53, legend_x + 32, legend_y + 73), fill="#e21a1a", width=5)
    draw.line((legend_x + 12, legend_y + 73, legend_x + 32, legend_y + 53), fill="#e21a1a", width=5)
    draw.text((legend_x + 43, legend_y + 46), "终点", font=legend_font, fill="#202020")
    image.save(output_png, format="PNG", optimize=True)


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert WGS84 GNSS data in Excel to local ENU coordinates")
    parser.add_argument("input", type=Path, nargs="?", default=Path("gnss_csv.xlsx"))
    parser.add_argument("--png", type=Path, default=Path("gnss_csv_enu_2d.png"))
    parser.add_argument("--xlsx", type=Path, default=Path("gnss_csv_enu.xlsx"))
    args = parser.parse_args()

    origin, valid_enu, east, north = convert_workbook(args.input, args.xlsx)
    plot_enu(args.png, origin, valid_enu, east, north)
    print(f"Origin (lon, lat, h): {origin}")
    print(f"Valid points: {len(valid_enu)}")
    print(f"East range: {min(p[0] for p in valid_enu):.4f} .. {max(p[0] for p in valid_enu):.4f} m")
    print(f"North range: {min(p[1] for p in valid_enu):.4f} .. {max(p[1] for p in valid_enu):.4f} m")
    print(f"Up range: {min(p[2] for p in valid_enu):.4f} .. {max(p[2] for p in valid_enu):.4f} m")
    print(f"PNG: {args.png.resolve()}")
    print(f"XLSX: {args.xlsx.resolve()}")


if __name__ == "__main__":
    main()
